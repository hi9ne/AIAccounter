"""
Reports API endpoints
Генерация PDF отчётов через APITemplate.io и CSV/Excel экспорт
"""
from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import text
from typing import Optional, Dict, Any, List
from datetime import date, datetime, timedelta
import httpx
import os
import io
import csv
import logging
import base64
import binascii

from app.database import get_db
from app.models.models import User, SavedReport
from app.schemas.report import ReportRequest, ReportResponse, ReportType
from app.utils.auth import get_current_user
from app.config import settings, Settings

logger = logging.getLogger(__name__)

router = APIRouter()

# APITemplate.io настройки
APITEMPLATE_BASE_URL = "https://rest.apitemplate.io/v2"


def _get_report_settings() -> Settings:
    """Load settings from environment/.env.

    Important: the server process may keep running while .env changes; reloading here
    allows updating APITemplate credentials without restarting the API.
    """

    try:
        return Settings()
    except Exception:
        # Fallback to already-loaded settings if reloading fails for any reason
        return settings  # type: ignore[return-value]


def _normalize_apitemplate_api_key(raw: str) -> str:
    """Normalize APITemplate API key.

    Some deployments store the key base64-encoded; if the value looks like base64
    and decodes into a printable, non-whitespace token, use decoded form.
    """

    value = (raw or "").strip().strip('"')
    if not value:
        return ""

    # If the key already contains separators, treat as raw key.
    if ":" in value:
        return value

    # Heuristic: some deployments store the key base64/base64url-encoded, sometimes without padding.
    # Try to decode, but only accept decoded tokens that look like an API key (non-whitespace, printable).
    candidate = value.replace("-", "+").replace("_", "/")
    # add padding if missing
    pad = (-len(candidate)) % 4
    if pad:
        candidate = candidate + ("=" * pad)

    try:
        decoded = base64.b64decode(candidate, validate=False)
        decoded_text = decoded.decode("utf-8", errors="strict").strip()
        if decoded_text and (" " not in decoded_text) and ("\n" not in decoded_text) and len(decoded_text) >= 16:
            return decoded_text
    except (binascii.Error, UnicodeDecodeError, ValueError):
        pass

    return value


def generate_ai_summary(stats: Dict, top_categories: List[Dict], period_days: int) -> str:
    """
    Генерирует простой текстовый AI комментарий для отчёта
    """
    comments = []
    
    total_income = stats.get('total_income', 0)
    total_expense = stats.get('total_expense', 0)
    balance = stats.get('balance', 0)
    expense_count = stats.get('expense_count', 0)
    
    # Баланс
    if balance > 0:
        savings_percent = round((balance / total_income * 100), 1) if total_income > 0 else 0
        comments.append(f"🎉 Отлично! Вы сэкономили {balance:,.0f} сом ({savings_percent}% от дохода).")
    elif balance < 0:
        comments.append(f"⚠️ Внимание! Расходы превысили доходы на {abs(balance):,.0f} сом.")
    else:
        comments.append("📊 Ваши доходы и расходы сбалансированы.")
    
    # Топ категория
    if top_categories:
        top_cat = top_categories[0]
        cat_name = top_cat.get('category', 'Неизвестно')
        cat_amount = top_cat.get('total_amount', 0)
        cat_percent = top_cat.get('percentage', 0)
        
        comments.append(f"💸 Больше всего вы потратили на «{cat_name}» — {cat_amount:,.0f} сом ({cat_percent}% всех расходов).")
        
        # Если есть вторая категория
        if len(top_categories) > 1:
            second_cat = top_categories[1]
            comments.append(f"📌 На втором месте — «{second_cat.get('category')}» ({second_cat.get('total_amount', 0):,.0f} сом).")
    
    # Средние траты
    if expense_count > 0 and period_days > 0:
        avg_daily = total_expense / period_days
        avg_per_transaction = total_expense / expense_count
        
        comments.append(f"📅 В среднем вы тратили {avg_daily:,.0f} сом в день.")
        
        if avg_per_transaction > 1000:
            comments.append(f"💡 Средняя покупка — {avg_per_transaction:,.0f} сом. Возможно, стоит обратить внимание на крупные траты.")
    
    # Количество транзакций
    if expense_count > period_days * 3:
        comments.append(f"🛒 За период вы совершили {expense_count} покупок — это довольно много! Мелкие траты могут накапливаться.")
    elif expense_count < period_days / 2:
        comments.append(f"✨ Всего {expense_count} покупок за период — отличный контроль над расходами!")
    
    return "\n".join(comments)


async def fetch_report_data(
    user_id: int,
    start_date: date,
    end_date: date,
    db: AsyncSession
) -> Dict[str, Any]:
    """
    Получить все данные для отчёта
    """
    # Базовая статистика (без зависимостей на SQL-функции)
    income_stats_result = await db.execute(
        text("""
            SELECT
                COALESCE(SUM(amount), 0) AS total_income,
                COUNT(*) AS income_count
            FROM income
            WHERE user_id = :user_id
              AND date >= :start_date
              AND date <= :end_date
              AND deleted_at IS NULL
        """),
        {"user_id": user_id, "start_date": start_date, "end_date": end_date}
    )
    total_income, income_count = income_stats_result.first() or (0, 0)

    expense_stats_result = await db.execute(
        text("""
            SELECT
                COALESCE(SUM(amount), 0) AS total_expense,
                COUNT(*) AS expense_count
            FROM expenses
            WHERE user_id = :user_id
              AND date >= :start_date
              AND date <= :end_date
              AND deleted_at IS NULL
        """),
        {"user_id": user_id, "start_date": start_date, "end_date": end_date}
    )
    total_expense, expense_count = expense_stats_result.first() or (0, 0)

    balance = float(total_income or 0) - float(total_expense or 0)

    # Топ категории расходов
    top_cat_result = await db.execute(
        text("""
            SELECT
                category,
                COALESCE(SUM(amount), 0) AS total_amount,
                COUNT(*) AS transaction_count
            FROM expenses
            WHERE user_id = :user_id
              AND date >= :start_date
              AND date <= :end_date
              AND deleted_at IS NULL
            GROUP BY category
            ORDER BY total_amount DESC
            LIMIT 10
        """),
        {"user_id": user_id, "start_date": start_date, "end_date": end_date}
    )
    top_categories = top_cat_result.fetchall()
    
    # Тренд баланса - используем прямой запрос вместо функции
    trend_query = text("""
        SELECT 
            COALESCE(i.date, e.date) as date,
            COALESCE(i.amount, 0) as income,
            COALESCE(e.amount, 0) as expense,
            COALESCE(i.amount, 0) - COALESCE(e.amount, 0) as balance
        FROM (
            SELECT date, SUM(amount) as amount
            FROM income
            WHERE user_id = :user_id
                AND date >= :start_date
                AND date <= :end_date
                AND deleted_at IS NULL
            GROUP BY date
        ) i
        FULL OUTER JOIN (
            SELECT date, SUM(amount) as amount
            FROM expenses
            WHERE user_id = :user_id
                AND date >= :start_date
                AND date <= :end_date
                AND deleted_at IS NULL
            GROUP BY date
        ) e ON i.date = e.date
        ORDER BY COALESCE(i.date, e.date)
    """)
    trend_result = await db.execute(trend_query, {
        "user_id": user_id,
        "start_date": start_date,
        "end_date": end_date
    })
    trend = trend_result.fetchall()
    
    # Все транзакции за период
    transactions_query = text("""
        SELECT 
            e.date,
            e.category,
            e.description,
            e.amount,
            'expense' as type
        FROM expenses e
        WHERE e.user_id = :user_id
            AND e.date >= :start_date
            AND e.date <= :end_date
            AND e.deleted_at IS NULL
        UNION ALL
        SELECT 
            i.date,
            i.category,
            i.description,
            i.amount,
            'income' as type
        FROM income i
        WHERE i.user_id = :user_id
            AND i.date >= :start_date
            AND i.date <= :end_date
            AND i.deleted_at IS NULL
        ORDER BY date DESC
    """)
    trans_result = await db.execute(transactions_query, {
        "user_id": user_id,
        "start_date": start_date,
        "end_date": end_date
    })
    transactions = trans_result.fetchall()
    
    return {
        "stats": {
            "total_income": float(total_income or 0),
            "total_expense": float(total_expense or 0),
            "balance": float(balance),
            "income_count": int(income_count or 0),
            "expense_count": int(expense_count or 0)
        },
        "top_categories": [
            {
                "category": cat[0],
                "total_amount": float(cat[1]),
                "transaction_count": int(cat[2]),
                "percentage": round((float(cat[1]) / float(total_expense) * 100), 1) if float(total_expense or 0) > 0 else 0.0
            }
            for cat in top_categories
        ],
        "balance_trend": [
            {
                "date": str(t[0]),
                "income": float(t[1]) if len(t) > 1 and t[1] is not None else 0.0,
                "expense": float(t[2]) if len(t) > 2 and t[2] is not None else 0.0,
                "balance": float(t[3]) if len(t) > 3 and t[3] is not None else 0.0
            }
            for t in trend
        ],
        "transactions": [
            {
                "date": str(t[0]),
                "category": t[1],
                "description": t[2],
                "amount": float(t[3]),
                "type": t[4]
            }
            for t in transactions
        ]
    }


async def generate_pdf_via_apitemplate(
    template_id: str,
    data: Dict[str, Any]
) -> str:
    """
    Генерация PDF через APITemplate.io
    Возвращает URL сгенерированного PDF
    """
    current_settings = _get_report_settings()
    api_key = _normalize_apitemplate_api_key(current_settings.APITEMPLATE_API_KEY or "")

    if not api_key:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="APITemplate.io API key not configured"
        )
    
    if not template_id:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Template ID not configured"
        )
    
    # APITemplate.io uses /create-pdf endpoint with template_id in query string
    url = f"{APITEMPLATE_BASE_URL}/create-pdf?template_id={template_id}"
    
    headers = {
        "X-API-KEY": api_key,
        "Content-Type": "application/json"
    }
    
    # APITemplate expects substitution variables inside the top-level "data" object.
    payload = {"data": data}
    
    logger.info(f"📄 Sending request to APITemplate.io")
    logger.info(f"   URL: {url}")
    logger.info(f"   Data keys: {list(data.keys())}")
    
    async with httpx.AsyncClient(timeout=60.0) as client:
        try:
            response = await client.post(url, json=payload, headers=headers)
            
            # Логируем ответ для отладки
            logger.info(f"   Response status: {response.status_code}")
            if response.status_code != 200:
                logger.error(f"   Response body: {response.text}")
            
            response.raise_for_status()
            
            result = response.json()
            
            if result.get("status") == "success":
                download_url = result.get("download_url") or result.get("download_url_pdf", "")
                logger.info(f"✅ PDF generated: {download_url}")
                return download_url
            else:
                raise HTTPException(
                    status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                    detail=f"PDF generation failed: {result.get('message', 'Unknown error')}"
                )
        except httpx.HTTPStatusError as e:
            body = ""
            try:
                body = (e.response.text or "").strip()
            except Exception:
                body = ""
            if body:
                body = body[:2000] + ("\n…(truncated)…" if len(body) > 2000 else "")

            logger.error(f"❌ APITemplate.io HTTP error: {str(e)}")
            if body:
                logger.error(f"   Response body: {body}")

            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=(
                    f"APITemplate.io request failed: {str(e)}"
                    + (f" | body: {body}" if body else "")
                ),
            )
        except httpx.HTTPError as e:
            logger.error(f"❌ APITemplate.io error: {str(e)}")
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"APITemplate.io request failed: {str(e)}",
            )


@router.post("/weekly", response_model=ReportResponse)
async def generate_weekly_report(
    week_start: Optional[date] = Query(None, description="Начало недели (по умолчанию - текущая неделя)"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """
    Генерация недельного отчёта (Weekly Report)
    """
    if not week_start:
        # Начало текущей недели (понедельник)
        today = date.today()
        week_start = today - timedelta(days=today.weekday())
    
    week_end = week_start + timedelta(days=6)
    
    # Получаем данные
    report_data = await fetch_report_data(current_user.user_id, week_start, week_end, db)
    
    # Генерируем AI комментарий
    ai_summary = generate_ai_summary(
        report_data['stats'],
        report_data['top_categories'],
        period_days=7
    )
    
    # Подготовка данных для шаблона
    template_data = {
        "report_type": "Недельный отчёт",
        "period": f"{week_start.strftime('%d.%m.%Y')} - {week_end.strftime('%d.%m.%Y')}",
        "generated_at": datetime.now().strftime("%d.%m.%Y %H:%M"),
        "user_name": f"{current_user.first_name} {current_user.last_name or ''}".strip(),
        "ai_summary": ai_summary,
        **report_data
    }

    # Генерация PDF
    current_settings = _get_report_settings()
    pdf_url = await generate_pdf_via_apitemplate(current_settings.WEEKLY_TEMPLATE_ID, template_data)

    # Сохраняем отчет в базу (чтобы он появился в miniapp)
    saved_report = SavedReport(
        user_id=current_user.user_id,
        report_type="weekly",
        title=f"Недельный отчёт {week_start.strftime('%d.%m.%Y')} - {week_end.strftime('%d.%m.%Y')}",
        period_start=week_start,
        period_end=week_end,
        pdf_url=pdf_url,
        format="pdf",
        report_data=template_data,
        expires_at=datetime.now() + timedelta(days=5),
    )
    db.add(saved_report)
    await db.commit()
    await db.refresh(saved_report)

    return {
        "report_type": "weekly",
        "start_date": week_start,
        "end_date": week_end,
        "pdf_url": pdf_url,
        "generated_at": datetime.now(),
    }


@router.post("/monthly", response_model=ReportResponse)
async def generate_monthly_report(
    year: Optional[int] = Query(None, description="Год"),
    month: Optional[int] = Query(None, ge=1, le=12, description="Месяц (1-12)"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """
    Генерация месячного отчёта (Monthly Report)
    """
    if not year or not month:
        # Текущий месяц
        today = date.today()
        year = today.year
        month = today.month
    
    # Первый и последний день месяца
    month_start = date(year, month, 1)
    if month == 12:
        month_end = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        month_end = date(year, month + 1, 1) - timedelta(days=1)
    
    # Получаем данные
    report_data = await fetch_report_data(current_user.user_id, month_start, month_end, db)

    # Бюджеты в текущей схеме хранятся помесячно (budgets.month/budget_amount) и не привязаны к категориям.
    # Для PDF-шаблона оставляем список бюджетов пустым, чтобы не падать на несовместимом SQL.
    budgets = []
    
    # Генерируем AI комментарий
    period_days = (month_end - month_start).days + 1
    ai_summary = generate_ai_summary(
        report_data['stats'],
        report_data['top_categories'],
        period_days=period_days
    )
    
    # Подготовка данных для шаблона
    month_names = ["", "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
                   "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"]
    
    template_data = {
        "report_type": "Месячный отчёт",
        "period": f"{month_names[month]} {year}",
        "generated_at": datetime.now().strftime("%d.%m.%Y %H:%M"),
        "user_name": f"{current_user.first_name} {current_user.last_name or ''}".strip(),
        "ai_summary": ai_summary,
        "budgets": [],
        **report_data
    }

    # Генерация PDF
    current_settings = _get_report_settings()
    pdf_url = await generate_pdf_via_apitemplate(current_settings.MONTHLY_TEMPLATE_ID, template_data)

    # Сохраняем отчет в базу (чтобы он появился в miniapp)
    saved_report = SavedReport(
        user_id=current_user.user_id,
        report_type="monthly",
        title=f"Месячный отчёт {month_names[month]} {year}",
        period_start=month_start,
        period_end=month_end,
        pdf_url=pdf_url,
        format="pdf",
        report_data=template_data,
        expires_at=datetime.now() + timedelta(days=5),
    )
    db.add(saved_report)
    await db.commit()
    await db.refresh(saved_report)

    return {
        "report_type": "monthly",
        "start_date": month_start,
        "end_date": month_end,
        "pdf_url": pdf_url,
        "generated_at": datetime.now(),
    }


@router.post("/period", response_model=ReportResponse)
async def generate_period_report(
    report_request: ReportRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """
    Генерация отчёта за произвольный период (Period Report)
    """
    # Получаем данные
    report_data = await fetch_report_data(
        current_user.user_id,
        report_request.start_date,
        report_request.end_date,
        db
    )
    
    # Генерируем AI комментарий
    period_days = (report_request.end_date - report_request.start_date).days + 1
    ai_summary = generate_ai_summary(
        report_data['stats'],
        report_data['top_categories'],
        period_days=period_days
    )
    
    # Подготовка данных для шаблона
    template_data = {
        "report_type": "Отчёт за период",
        "period": f"{report_request.start_date.strftime('%d.%m.%Y')} - {report_request.end_date.strftime('%d.%m.%Y')}",
        "generated_at": datetime.now().strftime("%d.%m.%Y %H:%M"),
        "user_name": f"{current_user.first_name} {current_user.last_name or ''}".strip(),
        "ai_summary": ai_summary,
        "include_transactions": report_request.include_transactions,
        **report_data
    }
    
    # Если не нужны транзакции, убираем их из данных
    if not report_request.include_transactions:
        template_data["transactions"] = []
    
    # Генерация PDF
    current_settings = _get_report_settings()
    pdf_url = await generate_pdf_via_apitemplate(current_settings.PERIOD_TEMPLATE_ID, template_data)
    
    # Сохраняем отчет в базу
    saved_report = SavedReport(
        user_id=current_user.user_id,
        report_type="period",
        title=f"Отчет за {report_request.start_date.strftime('%d.%m.%Y')} - {report_request.end_date.strftime('%d.%m.%Y')}",
        period_start=report_request.start_date,
        period_end=report_request.end_date,
        pdf_url=pdf_url,
        format="pdf",
        report_data=template_data,
        expires_at=datetime.now() + timedelta(days=5)  # APITemplate.io хранит 5 дней
    )
    db.add(saved_report)
    await db.commit()
    await db.refresh(saved_report)
    
    return {
        "report_type": "period",
        "user_id": current_user.user_id,
        "start_date": report_request.start_date,
        "end_date": report_request.end_date,
        "pdf_url": pdf_url,
        "generated_at": datetime.now(),
        "report_id": saved_report.id
    }


@router.get("/history")
async def get_report_history(
    limit: int = Query(10, ge=1, le=50, description="Количество отчётов"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """
    История сгенерированных отчётов
    """
    from sqlalchemy import select, desc
    
    # Получаем сохраненные отчеты
    query = select(SavedReport).where(
        SavedReport.user_id == current_user.user_id
    ).order_by(desc(SavedReport.created_at)).limit(limit)
    
    result = await db.execute(query)
    reports = result.scalars().all()
    
    return {
        "user_id": current_user.user_id,
        "reports": [
            {
                "id": report.id,
                "report_type": report.report_type,
                "title": report.title,
                "period_start": report.period_start.isoformat(),
                "period_end": report.period_end.isoformat(),
                "pdf_url": report.pdf_url,
                "format": report.format,
                "created_at": report.created_at.isoformat(),
                "expires_at": report.expires_at.isoformat() if report.expires_at else None
            }
            for report in reports
        ]
    }


@router.get("/export/csv")
async def export_transactions_csv(
    start_date: date = Query(..., description="Начальная дата"),
    end_date: date = Query(..., description="Конечная дата"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """
    Экспорт транзакций в CSV
    """
    # Получаем все транзакции
    transactions_query = text("""
        SELECT 
            e.date,
            'Расход' as type,
            e.category,
            e.description,
            e.amount,
            e.currency
        FROM expenses e
        WHERE e.user_id = :user_id
            AND e.date >= :start_date
            AND e.date <= :end_date
            AND e.deleted_at IS NULL
        UNION ALL
        SELECT 
            i.date,
            'Доход' as type,
            i.category,
            i.description,
            i.amount,
            i.currency
        FROM income i
        WHERE i.user_id = :user_id
            AND i.date >= :start_date
            AND i.date <= :end_date
            AND i.deleted_at IS NULL
        ORDER BY date DESC
    """)
    
    result = await db.execute(transactions_query, {
        "user_id": current_user.user_id,
        "start_date": start_date,
        "end_date": end_date
    })
    transactions = result.fetchall()
    
    # Сохраняем запись в историю
    saved_report = SavedReport(
        user_id=current_user.user_id,
        report_type="export",
        title=f"CSV экспорт {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}",
        period_start=start_date,
        period_end=end_date,
        pdf_url=None,
        format="csv",
        report_data={"start_date": start_date.isoformat(), "end_date": end_date.isoformat()},
        expires_at=datetime.now() + timedelta(days=30),
    )
    db.add(saved_report)
    await db.commit()
    await db.refresh(saved_report)
    
    # Создаем CSV в памяти
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Заголовки
    writer.writerow(['Дата', 'Тип', 'Категория', 'Описание', 'Сумма', 'Валюта'])
    
    # Данные
    for t in transactions:
        writer.writerow([
            t[0].strftime('%Y-%m-%d'),
            t[1],
            t[2],
            t[3] or '',
            f"{t[4]:.2f}",
            t[5] or 'KGS'
        ])
    
    # Возвращаем CSV
    output.seek(0)
    filename = f"transactions_{start_date}_{end_date}.csv"
    
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode('utf-8-sig')),  # utf-8-sig для правильного отображения в Excel
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )


@router.get("/export/excel")
async def export_transactions_excel(
    start_date: date = Query(..., description="Начальная дата"),
    end_date: date = Query(..., description="Конечная дата"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """
    Экспорт транзакций в Excel
    Требует установки: pip install openpyxl
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="openpyxl library not installed. Please install it with: pip install openpyxl"
        )
    
    # Получаем данные
    report_data = await fetch_report_data(current_user.user_id, start_date, end_date, db)
    
    # Сохраняем запись в историю
    saved_report = SavedReport(
        user_id=current_user.user_id,
        report_type="export",
        title=f"Excel отчёт {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}",
        period_start=start_date,
        period_end=end_date,
        pdf_url=None,
        format="xlsx",
        report_data={"start_date": start_date.isoformat(), "end_date": end_date.isoformat()},
        expires_at=datetime.now() + timedelta(days=30),
    )
    db.add(saved_report)
    await db.commit()
    await db.refresh(saved_report)
    
    # Создаем Excel файл
    wb = Workbook()
    
    # Лист 1: Сводка
    ws_summary = wb.active
    ws_summary.title = "Сводка"
    
    # Заголовок
    ws_summary['A1'] = f"Отчет за период {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}"
    ws_summary['A1'].font = Font(size=14, bold=True)
    
    # Статистика
    ws_summary['A3'] = "Показатель"
    ws_summary['B3'] = "Значение"
    ws_summary['A3'].font = Font(bold=True)
    ws_summary['B3'].font = Font(bold=True)
    
    stats = report_data['stats']
    ws_summary['A4'] = "Доходы"
    ws_summary['B4'] = stats['total_income']
    ws_summary['A5'] = "Расходы"
    ws_summary['B5'] = stats['total_expense']
    ws_summary['A6'] = "Баланс"
    ws_summary['B6'] = stats['balance']
    ws_summary['A7'] = "Количество доходов"
    ws_summary['B7'] = stats['income_count']
    ws_summary['A8'] = "Количество расходов"
    ws_summary['B8'] = stats['expense_count']
    
    # Лист 2: Транзакции
    ws_trans = wb.create_sheet("Транзакции")
    
    # Заголовки
    headers = ['Дата', 'Тип', 'Категория', 'Описание', 'Сумма']
    ws_trans.append(headers)
    
    # Стилизация заголовков
    for cell in ws_trans[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Данные
    for t in report_data['transactions']:
        ws_trans.append([
            t['date'],
            'Доход' if t['type'] == 'income' else 'Расход',
            t['category'],
            t['description'] or '',
            t['amount']
        ])
    
    # Лист 3: Топ категории
    ws_cat = wb.create_sheet("Категории")
    
    headers_cat = ['Категория', 'Сумма', 'Количество', 'Процент']
    ws_cat.append(headers_cat)
    
    for cell in ws_cat[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    for cat in report_data['top_categories']:
        ws_cat.append([
            cat['category'],
            cat['total_amount'],
            cat['transaction_count'],
            f"{cat['percentage']}%"
        ])
    
    # Сохраняем в память
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"report_{start_date}_{end_date}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )
