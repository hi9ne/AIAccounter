"""
Export API endpoints
Экспорт транзакций в CSV и Excel
"""
from fastapi import APIRouter, Depends, Query, Response
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import text
from typing import Optional
from datetime import date, datetime
import csv
import io

from app.database import get_db
from app.models.models import User
from app.utils.auth import get_current_user

router = APIRouter(prefix="/export", tags=["export"])


@router.get("/transactions")
async def export_transactions(
    format: str = Query("csv", description="Export format: csv or xlsx"),
    type: Optional[str] = Query(None, description="Filter by type: expense, income, or all"),
    start_date: Optional[date] = Query(None, description="Start date"),
    end_date: Optional[date] = Query(None, description="End date"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """
    Экспорт транзакций в CSV или Excel формате
    """
    # Базовый запрос
    query = """
        SELECT 
            date,
            type,
            category,
            amount,
            currency,
            description,
            created_at
        FROM (
            SELECT 
                date,
                'Расход' as type,
                category,
                amount,
                currency,
                description,
                created_at
            FROM expenses
            WHERE user_id = :user_id AND deleted_at IS NULL
            
            UNION ALL
            
            SELECT 
                date,
                'Доход' as type,
                category,
                amount,
                currency,
                description,
                created_at
            FROM income
            WHERE user_id = :user_id AND deleted_at IS NULL
        ) combined
        WHERE 1=1
    """
    
    params = {"user_id": current_user.user_id}
    
    # Фильтры
    if type == "expense":
        query = query.replace("WHERE 1=1", "WHERE type = 'Расход'")
    elif type == "income":
        query = query.replace("WHERE 1=1", "WHERE type = 'Доход'")
    
    if start_date:
        query += " AND date >= :start_date"
        params["start_date"] = start_date
    
    if end_date:
        query += " AND date <= :end_date"
        params["end_date"] = end_date
    
    query += " ORDER BY date DESC, created_at DESC"
    
    result = await db.execute(text(query), params)
    transactions = result.fetchall()
    
    # Формируем данные
    headers = ["Дата", "Тип", "Категория", "Сумма", "Валюта", "Описание"]
    
    rows = []
    for t in transactions:
        rows.append([
            t[0].strftime("%Y-%m-%d") if t[0] else "",
            t[1] or "",
            t[2] or "",
            float(t[3]) if t[3] else 0,
            t[4] or "KGS",
            t[5] or ""
        ])
    
    if format == "xlsx":
        return await export_xlsx(headers, rows, current_user.user_id)
    else:
        return export_csv(headers, rows, current_user.user_id)


def export_csv(headers: list, rows: list, user_id: int) -> StreamingResponse:
    """Экспорт в CSV"""
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';', quoting=csv.QUOTE_MINIMAL)
    
    # BOM для Excel
    output.write('\ufeff')
    
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    
    output.seek(0)
    
    filename = f"transactions_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Content-Type": "text/csv; charset=utf-8"
        }
    )


async def export_xlsx(headers: list, rows: list, user_id: int) -> StreamingResponse:
    """Экспорт в Excel (XLSX)"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        # Fallback на CSV если openpyxl не установлен
        return export_csv(headers, rows, user_id)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Транзакции"
    
    # Стили
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Заголовки
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Данные
    income_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    expense_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            
            # Подсветка по типу
            if row_data[1] == "Доход":
                cell.fill = income_fill
            elif row_data[1] == "Расход":
                cell.fill = expense_fill
            
            # Выравнивание суммы по правому краю
            if col_idx == 4:
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '#,##0.00'
    
    # Автоширина колонок
    for col in range(1, len(headers) + 1):
        max_length = len(str(headers[col - 1]))
        for row in range(2, len(rows) + 2):
            try:
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            except:
                pass
        ws.column_dimensions[get_column_letter(col)].width = min(max_length + 2, 50)
    
    # Закрепляем заголовок
    ws.freeze_panes = 'A2'
    
    # Сохраняем в буфер
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"transactions_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )


@router.get("/summary")
async def export_summary(
    format: str = Query("csv", description="Export format: csv or xlsx"),
    start_date: Optional[date] = Query(None, description="Start date"),
    end_date: Optional[date] = Query(None, description="End date"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """
    Экспорт сводки по категориям
    """
    # Сводка по категориям расходов
    query = """
        SELECT 
            category,
            COUNT(*) as count,
            SUM(amount) as total,
            AVG(amount) as average,
            currency
        FROM expenses
        WHERE user_id = :user_id 
        AND deleted_at IS NULL
    """
    
    params = {"user_id": current_user.user_id}
    
    if start_date:
        query += " AND date >= :start_date"
        params["start_date"] = start_date
    
    if end_date:
        query += " AND date <= :end_date"
        params["end_date"] = end_date
    
    query += " GROUP BY category, currency ORDER BY total DESC"
    
    result = await db.execute(text(query), params)
    categories = result.fetchall()
    
    headers = ["Категория", "Количество", "Всего", "Среднее", "Валюта"]
    
    rows = []
    for c in categories:
        rows.append([
            c[0] or "Без категории",
            int(c[1]) if c[1] else 0,
            float(c[2]) if c[2] else 0,
            round(float(c[3]), 2) if c[3] else 0,
            c[4] or "KGS"
        ])
    
    if format == "xlsx":
        return await export_xlsx(headers, rows, current_user.user_id)
    else:
        return export_csv(headers, rows, current_user.user_id)
